from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side

HEADER_FILL = PatternFill(fill_type='solid', fgColor='008000')
HEADER_FONT = Font(name='Arial', size=10, color='FFFFFF', bold=True)
DATA_FONT = Font(name='Calibri', size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def _apply_sheet_styles(worksheet) -> None:
    worksheet.freeze_panes = 'A2'

    for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
        for cell in row:
            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = DATA_FONT
            cell.border = THIN_BORDER

    for col_cells in worksheet.columns:
        max_len = 0
        column_letter = col_cells[0].column_letter
        for cell in col_cells:
            cell_length = len(str(cell.value)) if cell.value is not None else 0
            if cell_length > max_len:
                max_len = cell_length
        worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 80)


def write_reconciliation_excel(output_path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for sheet_name in sheets:
            worksheet = workbook[sheet_name]
            _apply_sheet_styles(worksheet)

    return output_path
